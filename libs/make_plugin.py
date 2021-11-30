from openpyxl import load_workbook
from pprint import pprint
import json

alpha_index = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]

def get_score(start_num, end_num, alpha, load_sheet):
    score_alpha = alpha_index[alpha_index.index(alpha) - 1]
    score = 0
    score_in = 0
    for i in range(start_num, end_num + 1):
        if str(load_sheet[f'{alpha}{i}'].value) == 'True':
            score_in = 1
            score += int(load_sheet[f'{score_alpha}{i}'].value)

    if score_in == 0:
        score = None

    return score


def get_score_owas(start_num, alpha, load_sheet):
    try:
        score = int(load_sheet[f'{alpha}{start_num}'].value)
    except:
        score = None
    return score


def write_plugin(exel_name,idx):

    plugin = {
        "twist": {"neck": 0, "back": 0, "elbow": 0, "wrist": 0},
        "support": {"body": 0, "elbow": 0, "chair_n": 0, "chair_y": 0, 'sit' : 0},
        "handler": {"best": 0, "good": 0, "bad": 0, "no": 0},
        "action": {"4up": 0, "bad": 0},
        "weight": {'weight' : 0, 'sud' : 0, '4up' : 0}
    }

    load_exel = load_workbook(exel_name, data_only=True)
    # for i, sheet_n in enumerate(load_exel.sheetnames):
    #     print(f"num->{i} ------ {sheet_n} ")
    try:
        sheet_num = idx
        # sheet_num = int(input("원하는 num = "))
    except:
        print("숫자를 입력하시오")

    alpha_index = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]

    sheet_name = load_exel.sheetnames[sheet_num]
    # print(sheet_name)
    load_sheet = load_exel[sheet_name]

    # twist

    if str(load_sheet['C25'].value) == 'True':
        plugin['twist']['back'] = 1

    if str(load_sheet['C33'].value) == 'True':
        plugin['twist']['neck'] = 1

    if str(load_sheet['J40'].value) == 'True':
        plugin['twist']['wrist'] = 1

    if str(load_sheet['J23'].value) == 'True':
        plugin['twist']['elbow'] = 1

    # support

    if str(load_sheet['J25'].value) == 'True':
        plugin['support']['elbow'] = 1

    if str(load_sheet['C19'].value) == 'True':
        plugin['support']['chair_n'] = 1

    if str(load_sheet['C18'].value) == 'True':
        plugin['support']['chair_y'] = 1

    if str(load_sheet['C68'].value) == 'True':
        plugin['support']['sit'] = 1

    # handler

    if str(load_sheet['J73'].value) == 'True':
        plugin['handler']['best'] = 1

    if str(load_sheet['J74'].value) == 'True':
        plugin['handler']['good'] = 1

    if str(load_sheet['J75'].value) == 'True':
        plugin['handler']['bad'] = 1

    if str(load_sheet['J76'].value) == 'True':
        plugin['handler']['no'] = 1

    # weigth

    if str(load_sheet['C75'].value) == 'True':
        plugin['weight']['weight'] = 0

    if str(load_sheet['C76'].value) == 'True':
        plugin['weight']['weight'] = 5

    if str(load_sheet['I104'].value) == '2':
        plugin['weight']['weight'] = 10

    if str(load_sheet['I104'].value) == '3':
        plugin['weight']['weight'] = 20

    if str(load_sheet['C78'].value) == 'True':
        plugin['weight']['sud'] = 1


    # action

    if str(load_sheet['C82'].value) == 'True':
        plugin['action']['4up'] = 1
        plugin['weight']['4up']
    if str(load_sheet['C83'].value) == 'True':
        plugin['action']['bad'] = 1
        plugin['weight']['bad'] = 1

    return plugin

